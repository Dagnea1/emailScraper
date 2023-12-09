import re
import requests
import openpyxl


# read url from input
original_url = input("Enter the website url: ") #http://www.erdajt.com/utils/guestbook/index.php -> some emails https://www.randomlists.com/email-addresses?qty=200 https://www.onlinedatagenerator.com/Home/fakeemailaddressgenerator

#email_regex
EMAIL_REGEX="[\w\.-]+@[\w\.-]+"

#create a CSV file
workbook = openpyxl.Workbook()
workbook.save("emails.xlsx")
workbook = openpyxl.load_workbook("emails.xlsx")

#reference to the sheet
worksheet = workbook.get_sheet_by_name('Sheet')

#Cell Number
cellNumber = 1


#Response from webpage
response = requests.get(original_url)

for re_match in re.findall(EMAIL_REGEX, response.text):
    #Cell reference
    cell = worksheet.cell(row = cellNumber, column = 1)
    cell.value = re_match
    print(re_match)
    cellNumber += 1

workbook.save("emails.xlsx")